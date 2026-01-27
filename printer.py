import asyncio
import sys
from msvcrt import getch

import pandas as pd
import datetime

import os
from enum import Enum

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side

from pysnmp.hlapi.v3arch.asyncio import (
    get_cmd, SnmpEngine, CommunityData, UdpTransportTarget,
    ContextData, ObjectType, ObjectIdentity
)

g_szPath_Data = None
g_szPath_Export = None
g_iPageCountMax = 31
class ParsePrinterStatus(Enum):
    SUCSEFULL = 1
    ERROR_PARCE = 2
    ERROR_TONER_INFO = 3

async def Printer_Start():
    AllPrintersData = []
    AllNotifyUsers = []
    global g_szPath_Data
    global g_iPageCountMax
    bNotify = False
    szNotifyMessage = False
    LoopBreaker = 0
    iLowValue = 0

    bParcePrintes = False
    bParceUsers = False
    bSettings = False
    with open(g_szPath_Data, 'r', encoding="utf-8") as file:
        for text in file:

            if -1 != text.find("[Settings]"):
                bParceUsers = False
                bParcePrintes = False
                bSettings = True
                continue

            if bNotify:
                if -1 != text.find("[NotifyUsers]"):
                    bParceUsers = True
                    bParcePrintes = False
                    bSettings = False
                    continue
            
            if -1 != text.find("[Printers]"):
                bParcePrintes = True
                bParceUsers = False
                bSettings = False
                continue
            
            if (bSettings):
                if -1 != text.find("NotifyUsers="):
                    clean_content = text.replace('\n','')
                    aData = clean_content.split('=')
                    if (aData[1] == '0'):
                        bNotify = False
                    if (aData[1] == '1'):
                        bNotify = True
                    continue

                if -1 != text.find("PagePerDays="):
                    clean_content = text.replace('\n','')
                    aData = clean_content.split('=')
                    g_iPageCountMax = int(aData[1])
                    continue

                if -1 != text.find("NotifyMessage="):
                    clean_content = text.replace('\n','')
                    aData = clean_content.split('=')
                    szNotifyMessage = aData[1]
                    continue

                if -1 != text.find("TonerLow%="):
                    clean_content = text.replace('\n','')
                    aData = clean_content.split('=')
                    iLowValue = int(aData[1])
                    continue


            
            if (bParceUsers):
                clean_content = text.replace('\n','')
                AllNotifyUsers.append(clean_content)
                continue


            if (bParcePrintes):
                clean_content = text.replace('\n','')
                aData = clean_content.split('=')
                AllPrintersData.append({
                'IP': aData[1],
                'Name': aData[0],
                'DriverName': "NO DATA",
                'PrintCount': "NO DATA",
                'ParseStatus': ParsePrinterStatus.ERROR_PARCE
                })

                LoopBreaker += 1
            # if (LoopBreaker > 5):
            #     break

    futures = [asyncio.create_task(GetTonersStatus(PrinterInfo['IP'])) for PrinterInfo in AllPrintersData]
    done, pending = await asyncio.wait(futures)
    
    for future in done:
        aData = future.result()
        try:
            for index, aPrinter in enumerate(AllPrintersData):
                if (aPrinter['IP'] == aData[0]['IP']):
                    AllPrintersData[index]['DriverName'] = aData[0]['DriverName']
                    AllPrintersData[index]['PrintCount'] = aData[0]['PrintCount']
                    aData = aData[1:]

                    AllPrintersData[index]['ParseStatus'] = ParsePrinterStatus.SUCSEFULL
                    AllPrintersData[index]['HasCacha'] = False
                    if (aData[0]['description'] == ""):
                        AllPrintersData[index]['ParseStatus'] = ParsePrinterStatus.ERROR_TONER_INFO
                    else:
                        AllPrintersData[index]['TonerData'] = aData[::-1]
        except Exception:
            pass

    global g_szPath_Script
    szCachePath = g_szPath_Script + '\\PrinterParceCache.txt'

    # Нет файла кеша
    if not os.path.exists(szCachePath):
        f = open(szCachePath, 'w')
        for index, aPrinter in enumerate(AllPrintersData):

            szString = f"IP={aPrinter['IP']}\n"
            f.write(szString)

            szString = f"PRINT_COUNT={aPrinter['PrintCount']}\n"
            f.write(szString)

            iTime = int(datetime.datetime.now().timetuple().tm_yday)
            szString = f"PRINT_PARCE_TIME={iTime}\n"
            f.write(szString)

            AllPrintersData[index]['PrintCount'] = 0
        f.close()
    else:
        aReWrite = []

        with open(szCachePath, 'r', encoding="utf-8") as file:
            IP = None
            iCount = None
            for text in file:
                if -1 != text.find("IP="):
                    clean_content = text.replace('\n','')
                    aData = clean_content.split('=')
                    IP = aData[1]
                    continue
                if -1 != text.find("PRINT_COUNT="):
                    clean_content = text.replace('\n','')
                    aData = clean_content.split('=')
                    iCount = int(aData[1])
                    continue
                
                if -1 != text.find("PRINT_PARCE_TIME="):
                    clean_content = text.replace('\n','')
                    aData = clean_content.split('=')
                    iDate = int(aData[1])
                    for index, aPrinter in enumerate(AllPrintersData):
                        if (aPrinter['IP'] == IP):
                            if aPrinter['ParseStatus'] == ParsePrinterStatus.ERROR_PARCE or AllPrintersData[index]['PrintCount'] < 1:
                                break
                            
                            iTimeDays = int(datetime.datetime.now().timetuple().tm_yday - iDate)
                            if (iTimeDays >= 1):
                                aReWrite.append(index)
                            else:
                                AllPrintersData[index]['PrintCount'] = int(AllPrintersData[index]['PrintCount']) - iCount
                            AllPrintersData[index]['HasCacha'] = True
                            continue

        if len(aReWrite) > 0:
            with open(szCachePath, 'r', encoding="utf-8") as file:
                lines = file.readlines()

                ReWriteIndex = None

                for index, line in enumerate(lines):
                    if -1 != line.find("IP="):
                        clean_content = line.replace('\n','')
                        aData = clean_content.split('=')
                        IP = aData[1]
                        bFind = False
                        for rewrite in aReWrite:
                            if AllPrintersData[rewrite]['IP'] == IP:
                                ReWriteIndex = rewrite
                                bFind = True
                                break
                        if bFind:
                            continue

                    if ReWriteIndex != None:
                        if -1 != line.find("PRINT_COUNT="):
                            lines[index] = f"PRINT_COUNT={AllPrintersData[ReWriteIndex]['PrintCount']}\n"
                            AllPrintersData[ReWriteIndex]['PrintCount'] = 0
                        if -1 != line.find("PRINT_PARCE_TIME="):
                            lines[index] = f"PRINT_PARCE_TIME={str(datetime.datetime.now().timetuple().tm_yday)}\n"
                            ReWriteIndex = None

            with open(szCachePath, 'w', encoding='utf-8') as rfile:
                rfile.writelines(lines)

            for index in aReWrite:
                AllPrintersData[index]['HasCacha'] = True

        with open(szCachePath, 'a', encoding="utf-8") as file:
            for index, aPrinter in enumerate(AllPrintersData):
                if aPrinter['ParseStatus'] == ParsePrinterStatus.ERROR_PARCE or AllPrintersData[index]['PrintCount'] < 1:
                    continue
                if AllPrintersData[index]['HasCacha']:
                    continue
                szString = f"IP={aPrinter['IP']}\n"
                file.write(szString)

                szString = f"PRINT_COUNT={aPrinter['PrintCount']}\n"
                file.write(szString)

                iTime = int(datetime.datetime.now().timetuple().tm_yday)
                szString = f"PRINT_PARCE_TIME={iTime}\n"
                file.write(szString)

                AllPrintersData[index]['PrintCount'] = 0
                
    
    PrintToExelData(AllPrintersData)
    if bNotify and len(AllNotifyUsers) > 0:
        bLow = False
        if (iLowValue > 0):
            for index, aPrinter in enumerate(AllPrintersData):
                if (AllPrintersData[index]['ParseStatus'] != ParsePrinterStatus.SUCSEFULL):
                    continue
                
                for ID in range(0, len(AllPrintersData[index]['TonerData'])):
                    iValue = int(int(float(aPrinter['TonerData'][ID]['level']) / float(aPrinter['TonerData'][ID]['max_capacity']) * 100))
                    if (iValue <= iLowValue):
                        bLow = True
                        break
                if (bLow):
                    break
        if (bLow):
            for NotifyUser in AllNotifyUsers:
                szCommand = f"msg * /server:{str(NotifyUser)} \"{str(szNotifyMessage)}\""
                os.system(szCommand)


async def GetTonersStatus(ip, community='public'):
    base_oid = '1.3.6.1.2.1.43.11.1.1'
    base_color_oid = '1.3.6.1.2.1.43.12.1.1.4.1.'

    base_level_oid = f'{base_oid}.9.1.'  # prtMarkerSuppliesLevel
    base_max_oid   = f'{base_oid}.8.1.'  # prtMarkerSuppliesMaxCapacity
    base_desc_oid  = f'{base_oid}.6.1.'  # prtMarkerSuppliesDescription
    name_oid = '1.3.6.1.2.1.43.5.1.1.16.1'
    print_oid = '1.3.6.1.2.1.43.10.2.1.4.1.1'

    aID = [1, 2, 3, 4]
    aData = [{'IP': ip}]
    try:
        target = await UdpTransportTarget.create((ip, 161), timeout=2, retries=3)  

        szName_printer = await get_cmd(
                SnmpEngine(),
                CommunityData(community, mpModel=0),
                target,
                ContextData(),
                ObjectType(ObjectIdentity(name_oid))
        )
        szName_printer = szName_printer[3][0][1]
        aData[0]['DriverName'] = str(szName_printer)

        iPrintCount = await get_cmd(
                SnmpEngine(),
                CommunityData(community, mpModel=0),
                target,
                ContextData(),
                ObjectType(ObjectIdentity(print_oid))
        )
        iPrintCount = int(iPrintCount[3][0][1])
        aData[0]['PrintCount'] = iPrintCount

        for i in range(0, 4):
            level_oid = base_level_oid + str(aID[i])
            # PrintCount
            level_result = await get_cmd(
                SnmpEngine(),
                CommunityData(community, mpModel=0),
                target,
                ContextData(),
                ObjectType(ObjectIdentity(level_oid))
            )
            #get_cmd() returns (errorIndication, errorStatus, errorIndex, varBinds), varBinds[0][0] is oid, varBinds[0][1] is returned value
            if (str(level_result[1]) != "0"):
                aData = aData[:-1]
                break
            # return
            level = int(level_result[3][0][1])
            
            max_oid = base_max_oid + str(aID[i])
            # Get max capacity
            max_result = await get_cmd(
                SnmpEngine(),
                CommunityData(community, mpModel=0),
                target,
                ContextData(),
                ObjectType(ObjectIdentity(max_oid))
            )
            max_capacity = int(max_result[3][0][1])

            desc_oid = base_desc_oid + str(aID[i])
            # Get description (optional)
            desc_result = await get_cmd(
                SnmpEngine(),
                CommunityData(community, mpModel=0),
                target,
                ContextData(),
                ObjectType(ObjectIdentity(desc_oid))
            )
            desc = str(desc_result[3][0][1])


            color_oid = base_color_oid + str(aID[i])
            # Get description (optional)
            color_result = await get_cmd(
                SnmpEngine(),
                CommunityData(community, mpModel=0),
                target,
                ContextData(),
                ObjectType(ObjectIdentity(color_oid))
            )
            color = str(color_result[3][0][1])
            aData.append({
            'description': desc,
            'level': level,
            'max_capacity': max_capacity,
            'color': color})
        return aData
    except Exception:
        pass

def PrintToExelData(AllPrintersData):
    global g_szPath_Script
    g_szExelPathRead = g_szPath_Script + '\\Template.xlsx'
    g_hExel = pd.read_excel(g_szExelPathRead, skiprows=0, index_col=0, dtype=str)
    g_hExel.replace('nan', '')
    g_hExel = g_hExel.fillna('')

    IP_colum = g_hExel.columns.get_loc('IP')
    Name_colum = g_hExel.columns.get_loc('Name')
    ParceStatus_colum = g_hExel.columns.get_loc('Parce Status')
    ParceTime_colum = g_hExel.columns.get_loc('Parce Time')
    PrintTimes_colum = 4
    Color1_colum = 5

    for index, aPrinter in enumerate(AllPrintersData):
        IPIndex = -1
        FreeIndex = -1

        for Tableindex in range(0, g_hExel.shape[0]):
            # Поиск по айпи
            if (aPrinter['IP'] == str(g_hExel.iat[Tableindex, IP_colum])):
                IPIndex = Tableindex
                break

        if IPIndex == -1:
            for Tableindex in range(0, g_hExel.shape[0]):
                if g_hExel.iat[Tableindex, IP_colum] == '':
                    FreeIndex = Tableindex
                    break

            if FreeIndex == -1:
                print('No Free index ')
                return

            IPIndex = FreeIndex

        g_hExel.iat[Tableindex, IP_colum] = aPrinter['IP']
        g_hExel.iat[Tableindex, Name_colum] = aPrinter['Name'] + '/' + aPrinter['DriverName']
        g_hExel.iat[Tableindex, ParceStatus_colum] = GetParceStatusToString(aPrinter['ParseStatus'])

        g_hExel.iat[Tableindex, PrintTimes_colum] = aPrinter['PrintCount']
        
        if (aPrinter['ParseStatus'] == ParsePrinterStatus.SUCSEFULL):
            g_hExel.iat[Tableindex, ParceTime_colum] = datetime.datetime.now().strftime("%H:%M")

        iColumModif = -2
        for ID in range(0, 4):
            iColumModif += 2
            if (aPrinter['ParseStatus'] == ParsePrinterStatus.SUCSEFULL):
                if ID < len(AllPrintersData[index]['TonerData']):
                    g_hExel.iat[Tableindex, Color1_colum + iColumModif] = f"{aPrinter['TonerData'][ID]['color']}:{aPrinter['TonerData'][ID]['description']} {aPrinter['TonerData'][ID]['level']}/{aPrinter['TonerData'][ID]['max_capacity']}"
                    g_hExel.iat[Tableindex, Color1_colum + iColumModif + 1] = int(int(float(aPrinter['TonerData'][ID]['level']) / float(aPrinter['TonerData'][ID]['max_capacity']) * 100))# + '%'
                    continue

            g_hExel.iat[Tableindex, Color1_colum + iColumModif] = '-'
            g_hExel.iat[Tableindex, Color1_colum + iColumModif + 1] = '-'

    global g_szPath_Export
    global g_iPageCountMax

    # есть файл эксель
    if os.path.exists(g_szPath_Export):
        
        wb = load_workbook(g_szPath_Export)

        g_szExelPathRead = g_szPath_Script + '\\Temp.xlsx'

        with pd.ExcelWriter(g_szExelPathRead, engine='openpyxl') as writer:
            iPageCount = 0
            szTableName = GetTableName()
            g_hExel.to_excel(writer, sheet_name=szTableName, index=True)
            for sheet_name in wb.sheetnames:
                if iPageCount >= g_iPageCountMax:
                    break

                if sheet_name != szTableName:
                    df = pd.read_excel(g_szPath_Export, sheet_name=sheet_name)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    iPageCount += 1

        os.system("attrib -r " + g_szPath_Export)
        os.replace(g_szExelPathRead, g_szPath_Export)
    else:
        g_hExel.to_excel(g_szPath_Export, sheet_name=GetTableName())

    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    wb = load_workbook(g_szPath_Export)
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]

        # Border
        for row in ws:
            for cell in row:
                if cell.value is not None:
                    cell.border = border_style
                    cell.alignment = Alignment(horizontal='left', vertical='center')
        
        aCenterCol = [1, 4, 5, 6]
        for iCol in aCenterCol:
            for row in ws.iter_rows(min_row=2, min_col = iCol, max_col=iCol):
                for cell in row:
                    if cell.value is not None:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        # cell.fill = PatternFill(start_color="FF00FF", end_color="FF00FF", fill_type='solid')

        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            
            for cell in column_cells:
                if cell.value:
                    
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length

            adjusted_width = max_length + 1
            ws.column_dimensions[column].width = adjusted_width

        for row in ws.iter_rows(min_row=2, min_col = 8):
            for cell in row:
                if cell.value is not None:
                    try:
                        fValue = float(cell.value)
                        if fValue < 15:
                            fValue = 0
                        if fValue > 75:
                            fValue = 100
                        szColor = GetColorByProccent(fValue)
                        cell.fill = PatternFill(start_color=szColor, end_color=szColor, fill_type='solid')
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                    except Exception: 
                        pass
    wb.save(g_szPath_Export)

    os.system("attrib +r " + g_szPath_Export)

def GetTableName():
    return str(datetime.datetime.now().strftime("%d.%m"))

def GetColorByProccent(iProccent):
    iRed = 255
    if iProccent > 50:
        iRed = int((1 - 2 * (iProccent - 50) / 100.0) * 255)

    iGreen = 255
    if iProccent <= 50:
        iGreen = int((2 * iProccent / 100.0) * 255)
    iBlue = 0
    return f"{((iRed & 0xFF) << 16) | ((iGreen & 0xFF) << 8) | (iBlue & 0xFF):06X}"

def GetParceStatusToString(iValue):
    if iValue == ParsePrinterStatus.SUCSEFULL:
        return '✅'
    if iValue == ParsePrinterStatus.ERROR_PARCE:
        return '☐'
    return '☒'

def main():
    szArgs = sys.argv[1:]
    global g_szPath_Data
    global g_szPath_Export
    global g_szPath_Script

    if len(szArgs) != 3:
        print("Bad arguments. Press any key to exit...")
        junk = getch()

        # g_szPath_Data = "C:\\Users\\a.vyushkov\\Desktop\\printer_data.ini"
        # g_szPath_Export = "C:\\Users\\a.vyushkov\\Desktop\\Toners.xlsx"
        # g_szPath_Script = "C:\\Users\\a.vyushkov\\Desktop"
        # asyncio.run(Printer_Start())

        return

    g_szPath_Data = str(szArgs[0])
    g_szPath_Export = str(szArgs[1])
    g_szPath_Script = str(szArgs[2])

    asyncio.run(Printer_Start())

if __name__ == "__main__":
    main()